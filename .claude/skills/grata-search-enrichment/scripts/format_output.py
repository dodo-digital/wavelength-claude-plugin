#!/usr/bin/env python3
"""Format enrichment results into xlsx + csv output files.

Usage: python3 format_output.py <industry> <output_dir>
Input: JSON enrichment data from stdin
Output: JSON to stdout with file paths created.

Expected stdin format:
{
  "companies": [
    {
      "force_rank": "H1",
      "fit_rating": "HIGH",
      "descriptor": "...",
      "rationale": "...",
      "company_name": "...",
      "revenue": "...",
      "employees": "...",
      "year_founded": "...",
      "hq": "...",
      "website": "...",
      "owner_signals": "...",
      ...original grata columns...
    }
  ]
}
"""

import csv
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}))
    sys.exit(1)


GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD = Font(bold=True)

# Enrichment columns always first
ENRICH_COLS = ["force_rank", "fit_rating", "descriptor", "rationale"]

# Shortlist columns
SHORTLIST_COLS = [
    "force_rank", "company_name", "descriptor", "rationale",
    "revenue", "employees", "year_founded", "hq", "website", "owner_signals",
]


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2


def color_row(ws, row_idx, fit_rating):
    fill = None
    if fit_rating == "HIGH":
        fill = GREEN_FILL
    elif fit_rating == "MEDIUM":
        fill = YELLOW_FILL
    if fill:
        for cell in ws[row_idx]:
            cell.fill = fill


def write_enriched_xlsx(companies, output_path):
    wb = Workbook()

    # Sheet 1: All companies
    ws = wb.active
    ws.title = "Enriched Companies"

    # Determine all columns: enrichment cols + all other keys
    other_keys = []
    if companies:
        all_keys = set()
        for c in companies:
            all_keys.update(c.keys())
        other_keys = sorted(all_keys - set(ENRICH_COLS))

    all_cols = ENRICH_COLS + other_keys

    # Header row
    for col_idx, header in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD

    # Data rows
    for row_idx, company in enumerate(companies, 2):
        for col_idx, header in enumerate(all_cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=company.get(header, ""))
        color_row(ws, row_idx, company.get("fit_rating", ""))

    auto_width(ws)

    # Sheet 2: Summary stats
    ws2 = wb.create_sheet("Summary")
    high = sum(1 for c in companies if c.get("fit_rating") == "HIGH")
    medium = sum(1 for c in companies if c.get("fit_rating") == "MEDIUM")
    low = sum(1 for c in companies if c.get("fit_rating") == "LOW")

    stats = [
        ("Total Companies", len(companies)),
        ("HIGH fit", high),
        ("MEDIUM fit", medium),
        ("LOW fit", low),
    ]
    for row_idx, (label, value) in enumerate(stats, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = BOLD
        ws2.cell(row=row_idx, column=2, value=value)

    auto_width(ws2)
    wb.save(output_path)


def write_shortlist_csv(companies, output_path):
    high_fit = [c for c in companies if c.get("fit_rating") == "HIGH"]
    high_fit.sort(key=lambda c: c.get("force_rank", "Z99"))

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SHORTLIST_COLS, extrasaction="ignore")
        writer.writeheader()
        for company in high_fit:
            writer.writerow(company)


def main():
    if len(sys.argv) != 3:
        print(json.dumps({"error": "Usage: format_output.py <industry> <output_dir>"}))
        sys.exit(1)

    industry = sys.argv[1]
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)

    data = json.loads(sys.stdin.read())
    companies = data.get("companies", [])

    today = date.today().isoformat()
    xlsx_path = output_dir / f"{industry}-{today}-enriched.xlsx"
    csv_path = output_dir / f"{industry}-{today}-shortlist.csv"

    write_enriched_xlsx(companies, str(xlsx_path))
    write_shortlist_csv(companies, str(csv_path))

    result = {
        "enriched_xlsx": str(xlsx_path),
        "shortlist_csv": str(csv_path),
        "total_companies": len(companies),
        "high_fit": sum(1 for c in companies if c.get("fit_rating") == "HIGH"),
        "medium_fit": sum(1 for c in companies if c.get("fit_rating") == "MEDIUM"),
        "low_fit": sum(1 for c in companies if c.get("fit_rating") == "LOW"),
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
