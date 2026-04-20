#!/usr/bin/env python3
"""Discover xlsx export structure. Pure I/O — no business logic.

Usage: python3 discover_export.py <path_to_xlsx>
Output: JSON to stdout with tab names, columns, row counts, sample values.
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}))
    sys.exit(1)


def discover(path: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"file": Path(path).name, "tabs": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=5, values_only=True))

        if not rows:
            result["tabs"].append({"name": sheet_name, "columns": [], "row_count": 0})
            continue

        headers = [str(h) if h is not None else f"_col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        # Count total rows (subtract header)
        row_count = ws.max_row - 1 if ws.max_row else 0

        columns = []
        for col_idx, header in enumerate(headers):
            samples = []
            for row in data_rows:
                if col_idx < len(row) and row[col_idx] is not None:
                    samples.append(str(row[col_idx])[:100])
            columns.append({
                "position": col_idx,
                "header": header,
                "samples": samples[:3],
            })

        result["tabs"].append({
            "name": sheet_name,
            "columns": columns,
            "row_count": row_count,
        })

    wb.close()
    return result


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(json.dumps({"error": "Usage: discover_export.py <path_to_xlsx>"}))
        sys.exit(1)

    path = sys.argv[1]
    if not Path(path).exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        sys.exit(1)

    print(json.dumps(discover(path), indent=2))
