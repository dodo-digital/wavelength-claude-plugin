#!/usr/bin/env python3
"""Extract owner-operator contacts from Grata export. Pure I/O + filtering — no business logic.

Usage: python3 extract_contacts.py <path_to_xlsx> [--industry <industry>]
Output: JSON to stdout with extracted contacts merged with company data.

This script handles:
  - Reading all three tabs (Companies, Top Exec, Other Exec)
  - Filtering to owner-operator titles only
  - Cross-referencing contacts to companies by company name
  - Extracting city/state from company headquarters
  - Deduplicating by email (keeps more senior title)
  - Populating LinkedIn with fallback URL if missing

Claude handles (NOT this script):
  - Generating business_model descriptors
  - Quality judgment calls
  - Scoring or rating
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}))
    sys.exit(1)

LINKEDIN_FALLBACK = "https://www.linkedin.com/in/denis-beslic-30bb6b25/"

# Owner-operator titles — case-insensitive matching
INCLUDE_TITLES = [
    "owner", "co-owner", "coowner",
    "president", "co-president",
    "founder", "co-founder", "cofounder",
    "ceo", "chief executive officer",
    "managing partner", "managing member",
    "principal",
]

EXCLUDE_TITLES = [
    "cfo", "chief financial officer",
    "cto", "chief technology officer",
    "coo", "chief operating officer",
    "cio", "chief information officer",
    "ciso", "chief information security officer",
    "cmo", "chief marketing officer",
    "vp", "vice president",
    "director",
    "manager",
    "board member", "advisor",
    "general manager",
    "former", "previous", "interim",
]

# Title seniority for dedup (lower = more senior)
TITLE_SENIORITY = {
    "owner": 1, "co-owner": 1,
    "founder": 2, "co-founder": 2,
    "ceo": 3, "chief executive officer": 3,
    "president": 4, "co-president": 4,
    "managing partner": 5, "managing member": 5,
    "principal": 6,
}


def is_owner_operator(title: str) -> bool:
    """Check if a title indicates an owner-operator."""
    if not title:
        return False
    lower = title.lower().strip()

    # Check exclusions first (catches "Former CEO", "VP of Operations", etc.)
    for exc in EXCLUDE_TITLES:
        if exc in lower and not any(inc in lower for inc in ["owner", "founder"]):
            return False

    # Check inclusions
    for inc in INCLUDE_TITLES:
        if inc in lower:
            return True

    return False


def title_seniority_score(title: str) -> int:
    """Return seniority score for dedup. Lower = more senior."""
    if not title:
        return 99
    lower = title.lower().strip()
    for key, score in TITLE_SENIORITY.items():
        if key in lower:
            return score
    return 99


def parse_location(location: str) -> tuple:
    """Parse 'City, State' or 'City, State, Country' into (city, state)."""
    if not location:
        return ("", "")
    parts = [p.strip() for p in location.split(",")]
    city = parts[0] if len(parts) > 0 else ""
    state = parts[1] if len(parts) > 1 else ""
    return (city, state)


def read_tab(wb, tab_name_candidates: list) -> tuple:
    """Find and read a tab by trying multiple name candidates.
    Returns (sheet_name, headers, rows) or (None, [], []).
    """
    for name in tab_name_candidates:
        for sheet_name in wb.sheetnames:
            if name.lower() in sheet_name.lower():
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return (sheet_name, [], [])
                headers = [str(h).strip() if h is not None else f"_col_{i}"
                           for i, h in enumerate(all_rows[0])]
                data = []
                for row in all_rows[1:]:
                    record = {}
                    for i, header in enumerate(headers):
                        val = row[i] if i < len(row) else None
                        record[header] = str(val).strip() if val is not None else ""
                    data.append(record)
                return (sheet_name, headers, data)
    return (None, [], [])


def find_column(headers: list, candidates: list) -> str | None:
    """Find a column header by trying multiple name candidates (case-insensitive)."""
    for candidate in candidates:
        for header in headers:
            if candidate.lower() == header.lower():
                return header
    # Partial match fallback
    for candidate in candidates:
        for header in headers:
            if candidate.lower() in header.lower():
                return header
    return None


def format_revenue(raw: str) -> str:
    """Convert raw revenue to human-readable format."""
    if not raw:
        return ""
    try:
        num = float(raw)
        if num >= 1_000_000:
            return f"${num / 1_000_000:.0f}M"
        elif num >= 1_000:
            return f"${num / 1_000:.0f}K"
        return f"${num:.0f}"
    except (ValueError, TypeError):
        return raw  # Already formatted or unparseable


def extract(path: str, industry: str = "") -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    stats = {
        "total_execs_found": 0,
        "owner_operators_kept": 0,
        "non_operators_dropped": 0,
        "duplicates_removed": 0,
        "missing_linkedin": 0,
        "missing_email": 0,
        "tabs_found": [],
    }

    # --- Read Companies tab ---
    co_name, co_headers, co_data = read_tab(wb, [
        "Companies", "Company Data", "Company List"
    ])
    if co_name:
        stats["tabs_found"].append(co_name)

    # Build company lookup by name
    co_name_col = find_column(co_headers, ["Name", "Company Name"])
    co_desc_col = find_column(co_headers, ["Description"])
    co_rev_col = find_column(co_headers, ["Revenue Estimate", "Revenue"])
    co_emp_col = find_column(co_headers, ["Employee Estimate", "Employees"])
    co_hq_col = find_column(co_headers, ["Headquarters", "HQ"])
    co_founded_col = find_column(co_headers, ["Year Founded", "Founded"])
    co_website_col = find_column(co_headers, ["Domain", "Website"])
    co_ownership_col = find_column(co_headers, ["Ownership", "Ownership Type"])
    co_naics_col = find_column(co_headers, ["NAICS 2", "NAICS"])

    company_lookup = {}
    for co in co_data:
        name = co.get(co_name_col, "") if co_name_col else ""
        if name:
            city, state = parse_location(co.get(co_hq_col, "") if co_hq_col else "")
            company_lookup[name.lower()] = {
                "company_name": name,
                "description": co.get(co_desc_col, "") if co_desc_col else "",
                "revenue": format_revenue(co.get(co_rev_col, "") if co_rev_col else ""),
                "employees": co.get(co_emp_col, "") if co_emp_col else "",
                "city": city,
                "state": state,
                "year_founded": co.get(co_founded_col, "") if co_founded_col else "",
                "website": co.get(co_website_col, "") if co_website_col else "",
                "ownership": co.get(co_ownership_col, "") if co_ownership_col else "",
                "naics": co.get(co_naics_col, "") if co_naics_col else "",
            }

    # --- Read executive tabs ---
    contacts = []

    for tab_candidates, tab_label in [
        (["Top Executive Contacts", "Top Executives", "Primary Contacts"], "top_exec"),
        (["Other Executive Contacts", "Other Executives", "Additional Contacts", "Secondary Contacts"], "other_exec"),
    ]:
        tab_name, headers, data = read_tab(wb, tab_candidates)
        if not tab_name:
            continue
        stats["tabs_found"].append(tab_name)

        # Find columns
        first_col = find_column(headers, ["First Name"])
        last_col = find_column(headers, ["Last Name"])
        company_col = find_column(headers, ["Company", "Company Name"])
        title_col = find_column(headers, ["Title", "Job Title"])
        email_col = find_column(headers, ["Work Email", "Email"])
        linkedin_col = find_column(headers, ["LinkedIn", "LinkedIn URL"])
        age_col = find_column(headers, ["Age"])

        for row in data:
            stats["total_execs_found"] += 1
            title = row.get(title_col, "") if title_col else ""

            if not is_owner_operator(title):
                stats["non_operators_dropped"] += 1
                continue

            first_name = row.get(first_col, "") if first_col else ""
            last_name = row.get(last_col, "") if last_col else ""
            company_name = row.get(company_col, "") if company_col else ""
            email = row.get(email_col, "") if email_col else ""
            linkedin = row.get(linkedin_col, "") if linkedin_col else ""
            age = row.get(age_col, "") if age_col else ""

            if not email:
                stats["missing_email"] += 1

            # Normalize LinkedIn
            if linkedin and not linkedin.startswith("http"):
                linkedin = f"https://{linkedin}"
            if not linkedin:
                linkedin = LINKEDIN_FALLBACK
                stats["missing_linkedin"] += 1

            # Get company data
            co_info = company_lookup.get(company_name.lower(), {})

            contacts.append({
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "company_name": company_name,
                "title": title,
                "industry": industry,
                "year_founded": co_info.get("year_founded", ""),
                "city": co_info.get("city", ""),
                "state": co_info.get("state", ""),
                "linkedin_profile": linkedin,
                "description": co_info.get("description", ""),
                "revenue": co_info.get("revenue", ""),
                "employees": co_info.get("employees", ""),
                "website": co_info.get("website", ""),
                "ownership": co_info.get("ownership", ""),
                "naics": co_info.get("naics", ""),
                "age": age,
                "source_tab": tab_label,
                "_title_seniority": title_seniority_score(title),
            })

    wb.close()

    # --- Deduplicate by email ---
    seen = {}
    deduped = []
    for contact in contacts:
        email_key = contact["email"].lower()
        if not email_key:
            deduped.append(contact)
            continue
        if email_key in seen:
            existing = seen[email_key]
            if contact["_title_seniority"] < existing["_title_seniority"]:
                # Replace with more senior title
                deduped = [c for c in deduped if c.get("email", "").lower() != email_key]
                deduped.append(contact)
                seen[email_key] = contact
                stats["duplicates_removed"] += 1
            else:
                stats["duplicates_removed"] += 1
        else:
            seen[email_key] = contact
            deduped.append(contact)

    # Clean up internal fields
    for contact in deduped:
        del contact["_title_seniority"]

    stats["owner_operators_kept"] = len(deduped)

    return {
        "contacts": deduped,
        "stats": stats,
        "industry": industry,
        "company_count": len(company_lookup),
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Extract owner-operator contacts from Grata export")
    parser.add_argument("path", help="Path to xlsx file")
    parser.add_argument("--industry", default="", help="Target industry (batch constant)")
    args = parser.parse_args()

    if not Path(args.path).exists():
        print(json.dumps({"error": f"File not found: {args.path}"}))
        sys.exit(1)

    result = extract(args.path, args.industry)
    print(json.dumps(result, indent=2))
